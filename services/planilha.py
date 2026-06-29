import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.receitanetbx import limpar_cnpj


def normalizar_texto(valor) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip().lower()

    substituicoes = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "õ": "o", "ô": "o",
        "ú": "u",
        "ç": "c",
    }

    for antigo, novo in substituicoes.items():
        texto = texto.replace(antigo, novo)

    return re.sub(r"[^a-z0-9]", "", texto)


def localizar_coluna_por_alias(ws, aliases: List[str], max_linhas: int = 20) -> Tuple[Optional[int], Optional[int]]:
    aliases_normalizados = {normalizar_texto(alias) for alias in aliases}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_linhas)):
        for cell in row:
            if normalizar_texto(cell.value) in aliases_normalizados:
                return cell.row, cell.column

    return None, None


def localizar_coluna_cnpj(ws) -> Tuple[int, int]:
    linha, coluna = localizar_coluna_por_alias(ws, ["cnpj"])

    if not linha or not coluna:
        raise ValueError("Não encontrei uma coluna chamada CNPJ nas primeiras 20 linhas da planilha.")

    return linha, coluna


def ler_planilha_entrada(caminho_arquivo: Path, ano_calendario: int, lote_id: str) -> Dict:
    wb = load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active

    linha_cabecalho, coluna_cnpj = localizar_coluna_cnpj(ws)
    _, coluna_codigo = localizar_coluna_por_alias(ws, [
        "codigo",
        "cod",
        "código",
        "codigo dominio",
        "cod dominio",
        "codi_emp"
    ])
    _, coluna_razao = localizar_coluna_por_alias(ws, [
        "razao social",
        "razão social",
        "razao",
        "razão",
        "razao social empresa",
        "razão social empresa",

        "nome",
        "nome empresa",
        "nome da empresa",
        "nome do cliente",
        "nome cliente",
        "cliente",

        "empresa",
        "empresas",
        "nome empresarial",
    ])

    registros = []
    vistos = set()
    ignorados = 0
    invalidos = 0

    for row_idx in range(linha_cabecalho + 1, ws.max_row + 1):
        cnpj_original = ws.cell(row=row_idx, column=coluna_cnpj).value
        cnpj = limpar_cnpj(cnpj_original)

        if not cnpj:
            ignorados += 1
            continue

        chave = (cnpj, ano_calendario)
        if chave in vistos:
            ignorados += 1
            continue

        vistos.add(chave)

        codigo = ws.cell(row=row_idx, column=coluna_codigo).value if coluna_codigo else None
        razao_social = ws.cell(row=row_idx, column=coluna_razao).value if coluna_razao else None

        status_inicial = "PENDENTE"
        observacao = None

        if len(cnpj) != 14:
            status_inicial = "CNPJ_INVALIDO"
            observacao = f"CNPJ inválido na linha {row_idx}: {cnpj_original}"
            invalidos += 1

        registros.append({
            "lote_id": lote_id,
            "linha_planilha": row_idx,
            "codigo": str(codigo).strip() if codigo is not None else None,
            "razao_social": str(razao_social).strip() if razao_social is not None else None,
            "cnpj_original": str(cnpj_original).strip() if cnpj_original is not None else None,
            "cnpj": cnpj,
            "ano_calendario": ano_calendario,
            "status": status_inicial,
            "observacao": observacao,
            "qtd_arquivos": 0,
            "ids_arquivos": [],
            "solicitado": "NÃO",
            "numero_pedido": None,
            "mensagem": None,
            "mensagem_solicitacao": None,
            "tentativas": 0,
            "created_at": datetime.now(),
            "updated_at": datetime.now(),
        })

    return {
        "linha_cabecalho": linha_cabecalho,
        "coluna_cnpj": get_column_letter(coluna_cnpj),
        "coluna_codigo": get_column_letter(coluna_codigo) if coluna_codigo else None,
        "coluna_razao_social": get_column_letter(coluna_razao) if coluna_razao else None,
        "registros": registros,
        "ignorados": ignorados,
        "invalidos": invalidos,
        "total_validos_unicos": len([r for r in registros if r["status"] != "CNPJ_INVALIDO"]),
        "total_registros": len(registros),
    }


def exportar_consultas_excel(consultas: List[Dict], caminho_saida: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    tipo_declaracao = consultas[0].get("tipo_declaracao", "ECD").upper() if consultas else "ECD"
    ws.title = f"Resultado {tipo_declaracao}"

    colunas = [
        ("Código", "codigo"),
        ("Razão Social", "razao_social"),
        ("CNPJ Original", "cnpj_original"),
        ("CNPJ Limpo", "cnpj"),
        ("Ano Consulta", "ano_calendario"),
        ("Status", "status"),
        ("Mensagem ReceitanetBX", "mensagem"),
        ("Qtd Arquivos", "qtd_arquivos"),
        ("IDs Arquivos", "ids_arquivos"),
        ("Solicitado", "solicitado"),
        ("Número Pedido", "numero_pedido"),
        ("Mensagem Solicitação", "mensagem_solicitacao"),
        ("Download Localizado", "download_localizado"),
        ("Nome Arquivo Baixado", "nome_arquivo_baixado"),
        ("Caminho Arquivo Baixado", "caminho_arquivo_baixado"),
        ("Tamanho Arquivo", "tamanho_arquivo_baixado"),
        ("Mensagem Download", "mensagem_download"),
        ("Data Download", "data_download"),
        ("Tentativas", "tentativas"),
        ("Observação", "observacao"),
        ("Data/Hora Consulta", "data_consulta"),
        ("Atualizado em", "updated_at"),
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (titulo, _) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, item in enumerate(consultas, start=2):
        for col_idx, (_, chave) in enumerate(colunas, start=1):
            valor = item.get(chave)

            if isinstance(valor, list):
                valor = " | ".join(str(v) for v in valor)
            elif isinstance(valor, datetime):
                valor = valor.strftime("%d/%m/%Y %H:%M:%S")

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col_idx in range(1, len(colunas) + 1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = 22

    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["G"].width = 55
    ws.column_dimensions["I"].width = 70
    ws.column_dimensions["L"].width = 45
    ws.column_dimensions["N"].width = 45

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida
